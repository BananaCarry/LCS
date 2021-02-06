from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from Table import table_gen
from itertools import count
from openpyxl.styles import Font,Color
from openpyxl.styles.alignment import Alignment

def convert_in_letter(it):
    for i in it:
        yield get_column_letter(i)

file_name = input("File name: ")
proposition = input("Proposition: ")

if ".xlsx" not in file_name:
    file_name += ".xlsx"

table = table_gen(proposition)

wb = Workbook()
sheet = wb.active
sheet.title = "Prop Table"

for row_index,row in zip(count(1), table):
    for column_letter,item in zip(convert_in_letter(count(1)), row):
        cell = sheet[column_letter + str(row_index)]
        cell.value = item

        font = Font()
        if item is True:
            font = Font(color = Color(rgb = "00158C8C"))
        elif item is False:
            font = Font(color = Color(rgb = "00991111"))
        cell.font = font

        if row_index == 1:
            cell.alignment = Alignment(horizontal = "center")

wb.save(file_name)